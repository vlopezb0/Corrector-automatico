from openpyxl import load_workbook
from openpyxl import Workbook


#Abrir accesos
wb = load_workbook(filename = 'C:\\Users\\Víctor\\Desktop\\TFM\\Excels\\Accesos.xlsx')

#Crear nuevo wb
wb2 = Workbook()
ws = wb2.active

sheet_ranges = wb['accesos']
inicio = -1

for i in range(613):
    actual = sheet_ranges['A'+str(i+2)].value
    hora_actual = sheet_ranges['B'+str(i+2)].value

    if inicio == -1:
        inicio = actual
        hora_antes = hora_actual
        # ws.append([actual,hora_actual])
        continue

    if actual != inicio:
        hora_final = sheet_ranges['B'+str(i+1)].value
        ws.append([inicio,hora_antes,hora_final])
        inicio = actual 
        hora_antes = hora_actual
        # ws.append([actual,hora_actual])

# ws.append([inicio,hora_antes,hora_final])


wb2.save("AccesosMOD.xlsx")
    # print(str(sheet_ranges['A'+str(i+2)].value), str(sheet_ranges['B'+str(i+2)].value))
